# -*- coding: utf-8 -*-
"""
สกัดข้อมูลจาก PDF งบทดลอง → Excel พร้อมสรุปหมวด
"""
import pdfplumber
import re
import sys
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

PDF_PATH = r'C:\Users\manat\OneDrive\ระบบงาน 4\งบการเงินปี 2568\งบทดลองสช.pdf'
OUT_PATH = r'C:\Users\manat\OneDrive\ระบบงาน 4\งบการเงินปี 2568\งบทดลองสช.xlsx'

CATEGORIES = {
    '1': 'สินทรัพย์',
    '2': 'หนี้สิน',
    '3': 'ทุน / ส่วนของเจ้าของ',
    '4': 'รายได้',
    '5': 'ค่าใช้จ่าย',
    '6': 'ค่าใช้จ่าย',
}

def parse_row(line):
    """
    Row format (typical):
    <ACCT_CODE> <THAI_NAME-ignored> <Fwd_DR> <Fwd_CR> <Curr_DR> <Curr_CR> <End_DR> <End_CR>
    We want: acct_code + last 2 numbers (End_DR, End_CR)
    Numbers can be 0.00 or 123,456.78 (with thousand-sep commas)
    """
    # Find all numbers
    num_pat = r'-?[\d,]+\.\d{2}'
    nums = re.findall(num_pat, line)
    if len(nums) < 6:
        return None
    # Account code — first 4+ digit token at start (may have extra like '1111(00')
    m = re.match(r'^\s*(\d{4,6})', line)
    if not m:
        return None
    code = m.group(1)
    # Ending DR/CR = last 2 numbers
    try:
        end_dr = float(nums[-2].replace(',', ''))
        end_cr = float(nums[-1].replace(',', ''))
    except ValueError:
        return None
    return code, end_dr, end_cr

def extract_all_rows():
    rows = []
    with pdfplumber.open(PDF_PATH) as pdf:
        for page in pdf.pages:
            txt = page.extract_text() or ""
            for line in txt.split('\n'):
                r = parse_row(line)
                if r:
                    rows.append(r)
    return rows

def main():
    rows = extract_all_rows()
    print(f"Extracted {len(rows)} rows")

    # No dedup — same 4-digit code represents many sub-accounts (distinguished by Thai name)
    # Number each occurrence
    rows.sort(key=lambda x: x[0])
    counter = defaultdict(int)
    numbered = []
    for code, dr, cr in rows:
        counter[code] += 1
        label = code if counter[code] == 1 else f"{code}-{counter[code]}"
        numbered.append((label, code, dr, cr))
    rows = numbered

    # Group by first digit
    cat_totals = defaultdict(lambda: [0.0, 0.0])
    for label, code, dr, cr in rows:
        first = code[0]
        cat_totals[first][0] += dr
        cat_totals[first][1] += cr

    # Build Excel
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "รายละเอียด"

    thin = Side(border_style="thin", color="888888")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    cat_fill = PatternFill("solid", fgColor="FFE699")
    cat_font = Font(bold=True, size=11)

    # Sheet 1: รายละเอียด
    headers = ["รหัสบัญชี", "หมวด", "ยอดเดบิตคงเหลือ", "ยอดเครดิตคงเหลือ"]
    ws1.append(headers)
    for col in range(1, 5):
        c = ws1.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    current_cat = None
    for label, code, dr, cr in rows:
        first = code[0]
        cat = CATEGORIES.get(first, f"หมวด {first}")
        if current_cat != first:
            # Insert subtotal before new category (except first)
            if current_cat is not None:
                tdr, tcr = cat_totals[current_cat]
                ws1.append([f"รวม {CATEGORIES.get(current_cat, current_cat)}", "", tdr, tcr])
                r = ws1.max_row
                for col in range(1, 5):
                    ws1.cell(row=r, column=col).fill = cat_fill
                    ws1.cell(row=r, column=col).font = cat_font
                    ws1.cell(row=r, column=col).border = border
            current_cat = first
        ws1.append([label, cat, dr, cr])
        r = ws1.max_row
        for col in range(1, 5):
            ws1.cell(row=r, column=col).border = border
        ws1.cell(row=r, column=3).number_format = '#,##0.00'
        ws1.cell(row=r, column=4).number_format = '#,##0.00'

    # Last subtotal
    if current_cat is not None:
        tdr, tcr = cat_totals[current_cat]
        ws1.append([f"รวม {CATEGORIES.get(current_cat, current_cat)}", "", tdr, tcr])
        r = ws1.max_row
        for col in range(1, 5):
            ws1.cell(row=r, column=col).fill = cat_fill
            ws1.cell(row=r, column=col).font = cat_font
            ws1.cell(row=r, column=col).border = border
        ws1.cell(row=r, column=3).number_format = '#,##0.00'
        ws1.cell(row=r, column=4).number_format = '#,##0.00'

    # Grand total
    gdr = sum(v[0] for v in cat_totals.values())
    gcr = sum(v[1] for v in cat_totals.values())
    ws1.append(["รวมทั้งสิ้น", "", gdr, gcr])
    r = ws1.max_row
    for col in range(1, 5):
        ws1.cell(row=r, column=col).fill = PatternFill("solid", fgColor="70AD47")
        ws1.cell(row=r, column=col).font = Font(bold=True, color="FFFFFF", size=12)
        ws1.cell(row=r, column=col).border = border
    ws1.cell(row=r, column=3).number_format = '#,##0.00'
    ws1.cell(row=r, column=4).number_format = '#,##0.00'

    ws1.column_dimensions['A'].width = 14
    ws1.column_dimensions['B'].width = 28
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 20

    # Sheet 2: สรุปหมวด
    ws2 = wb.create_sheet("สรุปหมวด")
    ws2.append(["หมวดบัญชี", "เลขขึ้นต้น", "ยอดเดบิตรวม", "ยอดเครดิตรวม", "ยอดสุทธิ"])
    for col in range(1, 6):
        c = ws2.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    # Consolidate 5+6 into single ค่าใช้จ่าย
    display = [
        ('สินทรัพย์', ['1']),
        ('หนี้สิน', ['2']),
        ('ทุน / ส่วนของเจ้าของ', ['3']),
        ('รายได้', ['4']),
        ('ค่าใช้จ่าย', ['5', '6']),
    ]
    for name, keys in display:
        tdr = sum(cat_totals[k][0] for k in keys if k in cat_totals)
        tcr = sum(cat_totals[k][1] for k in keys if k in cat_totals)
        net = tdr - tcr
        ws2.append([name, ','.join(keys), tdr, tcr, net])
        r = ws2.max_row
        for col in range(1, 6):
            ws2.cell(row=r, column=col).border = border
        for col in (3, 4, 5):
            ws2.cell(row=r, column=col).number_format = '#,##0.00'

    # Grand total
    gdr = sum(v[0] for v in cat_totals.values())
    gcr = sum(v[1] for v in cat_totals.values())
    ws2.append(["รวมทั้งสิ้น", "", gdr, gcr, gdr - gcr])
    r = ws2.max_row
    for col in range(1, 6):
        ws2.cell(row=r, column=col).fill = PatternFill("solid", fgColor="70AD47")
        ws2.cell(row=r, column=col).font = Font(bold=True, color="FFFFFF", size=12)
        ws2.cell(row=r, column=col).border = border
    for col in (3, 4, 5):
        ws2.cell(row=r, column=col).number_format = '#,##0.00'

    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 20

    wb.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")
    print()
    print("=== สรุปหมวด ===")
    for name, keys in display:
        tdr = sum(cat_totals[k][0] for k in keys if k in cat_totals)
        tcr = sum(cat_totals[k][1] for k in keys if k in cat_totals)
        print(f"{name:30s}  DR={tdr:>20,.2f}  CR={tcr:>20,.2f}  Net={tdr-tcr:>20,.2f}")
    print(f"{'รวม':30s}  DR={gdr:>20,.2f}  CR={gcr:>20,.2f}")

if __name__ == '__main__':
    main()
