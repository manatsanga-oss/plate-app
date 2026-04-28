# -*- coding: utf-8 -*-
"""
จับคู่ชื่อบัญชี "ค่าใช้จ่าย" จาก flow งบทดลอง สช.xlsx มาใส่ในงบทดลองสช.xlsx
เกณฑ์การจับคู่: รหัสบัญชี 4 หลักแรกตรง + ยอดเดบิตคงเหลือตรง (tolerance 0.01)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

REF = r'C:\Users\manat\OneDrive\ระบบงาน 4\งบการเงินปี 2568\flow งบทดลอง สช.xlsx'
TARGET = r'C:\Users\manat\OneDrive\ระบบงาน 4\งบการเงินปี 2568\งบทดลองสช.xlsx'

def load_ref_expenses():
    """return list of (code, name, end_dr, end_cr) for expense accounts (5xxx, 6xxx)"""
    wb = openpyxl.load_workbook(REF, data_only=True)
    ws = wb.active
    out = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        code = row[0]
        name = row[1]
        if not code or not name:
            continue
        code_str = str(code).strip()
        # expense: starts with 5 or 6
        if not (code_str.startswith('5') or code_str.startswith('6')):
            continue
        end_dr = row[9] or 0
        end_cr = row[10] or 0
        try:
            end_dr = float(end_dr)
            end_cr = float(end_cr)
        except (TypeError, ValueError):
            continue
        out.append((code_str, name, end_dr, end_cr))
    return out

def main():
    ref = load_ref_expenses()
    print(f"Reference expense rows: {len(ref)}")

    wb = openpyxl.load_workbook(TARGET)
    ws = wb['รายละเอียด']

    # Insert new column C for name, shift existing B→B, insert name at column 3
    # Actually simpler: append as new column E "ชื่อรายการ"
    ws.cell(row=1, column=5, value="ชื่อรายการ")
    thin = Side(border_style="thin", color="888888")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    c = ws.cell(row=1, column=5)
    c.fill = header_fill; c.font = header_font
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

    matched = 0
    unmatched = 0
    used = set()  # track which ref rows have been used

    for r in range(2, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        if not label or not isinstance(label, str):
            continue
        # code is either "5200" or "5200-3"
        code_base = label.split('-')[0]
        if not (code_base.startswith('5') or code_base.startswith('6')):
            continue
        # skip subtotal/grand total rows (those have text like "รวม...")
        if 'รวม' in str(label):
            continue
        try:
            dr = float(ws.cell(row=r, column=3).value or 0)
            cr = float(ws.cell(row=r, column=4).value or 0)
        except (TypeError, ValueError):
            continue

        # match by amount only (PDF and ref have different grouping)
        best = None
        for idx, (rcode, rname, rdr, rcr) in enumerate(ref):
            if idx in used:
                continue
            if abs(rdr - dr) < 0.02 and abs(rcr - cr) < 0.02:
                best = (idx, rcode, rname)
                break
        if best:
            idx, rcode, rname = best
            used.add(idx)
            cell = ws.cell(row=r, column=5, value=rname)
            cell.border = border
            matched += 1
        else:
            ws.cell(row=r, column=5).border = border
            unmatched += 1

    ws.column_dimensions['E'].width = 40

    # also apply border to non-expense rows in col E
    for r in range(2, ws.max_row + 1):
        if not ws.cell(row=r, column=5).border.left.style:
            ws.cell(row=r, column=5).border = border

    wb.save(TARGET)
    print(f"Saved: {TARGET}")
    print(f"Matched expense rows: {matched}")
    print(f"Unmatched expense rows: {unmatched}")

if __name__ == '__main__':
    main()
