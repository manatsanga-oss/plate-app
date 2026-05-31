import pandas as pd
from openpyxl import load_workbook
import json
import zipfile
import xml.etree.ElementTree as ET

path = r'C:\Users\manat\OneDrive\ระบบเงินเดือน\ค่าคอมมิชั้นV.2.xlsx'

# 1) List sheets
wb = load_workbook(path, data_only=False)
print("=== SHEETS ===")
for s in wb.sheetnames:
    sh = wb[s]
    print(f"  {s}: {sh.max_row} rows x {sh.max_column} cols")

# 2) Read each sheet as preview
print("\n=== SHEET PREVIEWS (first 10 rows) ===")
all_sheets = pd.read_excel(path, sheet_name=None, header=None, nrows=15)
for name, df in all_sheets.items():
    print(f"\n--- {name} ---")
    print(df.to_string())

# 3) Power Query / Connections (Power Query stored in xl/connections.xml or customXml)
print("\n=== POWER QUERY EXTRACTION ===")
with zipfile.ZipFile(path, 'r') as z:
    names = z.namelist()
    pq_files = [n for n in names if 'customXml' in n or 'connection' in n.lower() or 'queryTable' in n.lower() or 'dataModel' in n.lower()]
    print("Files of interest:")
    for n in pq_files:
        print(f"  {n}")

    # ลองอ่าน customXml/item*.xml (Power Query มักอยู่ที่นี่)
    for n in names:
        if n.startswith('customXml/item') and n.endswith('.xml'):
            try:
                with z.open(n) as f:
                    content = f.read().decode('utf-8', errors='ignore')
                    if 'Mashup' in content or 'M' in content[:500]:
                        print(f"\n--- {n} (first 2000 chars) ---")
                        print(content[:2000])
            except Exception as e:
                print(f"  Error reading {n}: {e}")
